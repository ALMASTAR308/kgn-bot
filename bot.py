import os
import sqlite3
import io
import threading
from flask import Flask
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== FLASK KEEP-ALIVE SERVER ====================
web_app = Flask(__name__)

@web_app.route('/')
def home():
    return "KGN Referral Bot is running 24/7!"

def run_web():
    port = int(os.environ.get("PORT", 8080))
    web_app.run(host='0.0.0.0', port=port)

# ==================== BOT CONFIGURATION ====================
TOKEN = "8941166511:AAGMpCIaO91gRRCdMzgB-_NKTtpu9PJQbE0"
ADMIN_ID = 8471139505
BOT_USERNAME = "KGN_ref_bot"

PRESALE_URL = "https://tools.smithii.io/launch/solana/Kognitron-KGN-Presale"
WEBSITE_URL = "https://kognitronkgn.com/"
ANNOUNCEMENTS_CHANNEL = "https://t.me/KognitronKGN"
COMMUNITY_CHAT = "https://t.me/kognitronChat"
DISCORD_URL = "https://discord.gg/UvGeH8XVe"
TWITTER_PROJECT = "https://x.com/KognitronKGN"
TWITTER_FOUNDER = "https://x.com/mrmonzer7"
CONTACT_EMAIL = "admin@kognitronkgn.com"

# Campaign Constants
USER_MAX_CAP = 200.0
TOTAL_POOL_CAP = 500000.0

conn = sqlite3.connect("referrals.db", check_same_thread=False)
cursor = conn.cursor()
cursor.execute("""
CREATE TABLE IF NOT EXISTS users (
    user_id INTEGER PRIMARY KEY,
    username TEXT,
    referred_by INTEGER,
    wallet_address TEXT,
    kgn_balance REAL DEFAULT 0
)
""")
try:
    cursor.execute("ALTER TABLE users ADD COLUMN username TEXT")
except sqlite3.OperationalError:
    pass
conn.commit()

def get_total_distributed():
    cursor.execute("SELECT SUM(kgn_balance) FROM users")
    result = cursor.fetchone()[0]
    return float(result) if result else 0.0

def get_leaderboard_text():
    cursor.execute("""
    SELECT 
        u.user_id,
        u.kgn_balance,
        COUNT(r.user_id) as ref_count,
        u.wallet_address
    FROM users u
    LEFT JOIN users r ON u.user_id = r.referred_by
    GROUP BY u.user_id
    ORDER BY u.kgn_balance DESC, ref_count DESC
    LIMIT 10
    """)
    leaders = cursor.fetchall()
    
    if not leaders or (len(leaders) == 1 and leaders[0][1] == 0 and leaders[0][2] == 0):
        return "🏆 **KGN Top 10 Leaderboard**\n\nNo active leaders with points or referrals yet!"

    text = "🏆 **Kognitron AI ($KGN) Top 10 Leaderboard**\n\n"
    for rank, row in enumerate(leaders, start=1):
        uid, bal, refs, wallet_addr = row
        medal = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else f"#{rank}"
        
        if wallet_addr and len(wallet_addr) > 8:
            display_id = f"`{wallet_addr[:4]}...{wallet_addr[-4:]}`"
        else:
            display_id = f"User `{str(uid)[-4:]}`"

        text += f"{medal} {display_id} — **{bal:.1f} $KGN** ({refs} referrals)\n"
    
    return text

def get_official_links_keyboard(is_admin=False):
    keyboard = [
        [InlineKeyboardButton("🛒 Buy Presale (Smithii)", url=PRESALE_URL)],
        [InlineKeyboardButton("🏆 Leaderboard (Top 10)", callback_data="show_leaderboard")],
        [
            InlineKeyboardButton("🌐 Website", url=WEBSITE_URL),
            InlineKeyboardButton("🐦 Project X", url=TWITTER_PROJECT)
        ],
        [
            InlineKeyboardButton("📢 Announcements", url=ANNOUNCEMENTS_CHANNEL),
            InlineKeyboardButton("💬 Telegram Chat", url=COMMUNITY_CHAT)
        ],
        [
            InlineKeyboardButton("👾 Discord", url=DISCORD_URL),
            InlineKeyboardButton("👤 Founder X", url=TWITTER_FOUNDER)
        ]
    ]
    if is_admin:
        keyboard.append([InlineKeyboardButton("📊 Export Airdrop Data (Admin Only)", callback_data="export_data")])
    return InlineKeyboardMarkup(keyboard)

async def generate_and_send_export(chat_id: int, context: ContextTypes.DEFAULT_TYPE):
    cursor.execute("""
    SELECT
        COALESCE(u.username, 'No Username') as username,
        u.user_id,
        COALESCE(u.wallet_address, 'Not Registered') as wallet,
        u.kgn_balance,
        COUNT(r.user_id) as referrals_count
    FROM users u
    LEFT JOIN users r ON u.user_id = r.referred_by
    GROUP BY u.user_id
    ORDER BY u.kgn_balance DESC, referrals_count DESC
    """)
    rows = cursor.fetchall()

    if not rows:
        await context.bot.send_message(chat_id=chat_id, text="⚠️ No user data found in the database yet.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "KGN Airdrop List"

    headers = ["الترتيب (Rank)", "اسم المستخدم (Username)", "User ID", "محفظة سولانا (Solana Wallet)", "رصيد المكافآت ($KGN)", "عدد الإحالات (Referrals)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for rank, row in enumerate(rows, start=1):
        username, uid, wallet, bal, refs = row
        ws.append([f"#{rank}", username, str(uid), wallet, bal, refs])
        
        row_idx = rank + 1
        for col_idx in range(1, 7):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = center_align
            c.border = thin_border
            if rank % 2 == 0:
                c.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    output.name = "kgn_airdrop_data.xlsx"

    total_dist = get_total_distributed()
    total_users = len(rows)

    caption = (
        "📊 **Kognitron AI ($KGN) Airdrop Master Sheet**\n\n"
        f"• **Total Registered Users:** `{total_users}`\n"
        f"• **Total Rewards Allocated:** `{total_dist:,.1f} / {TOTAL_POOL_CAP:,.0f} $KGN`\n"
        f"• **Individual Cap:** `{int(USER_MAX_CAP)} $KGN`\n\n"
        "📁 File is formatted as **Excel (.xlsx)** with organized columns."
    )

    await context.bot.send_document(
        chat_id=chat_id,
        document=output,
        filename="kgn_airdrop_data.xlsx",
        caption=caption,
        parse_mode="Markdown"
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        await update.message.reply_text(
            f"👋 To access your referral dashboard and register your wallet, start a private chat with me: @{BOT_USERNAME}"
        )
        return

    user_id = update.effective_user.id
    username = f"@{update.effective_user.username}" if update.effective_user.username else update.effective_user.first_name
    args = context.args

    cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
    user = cursor.fetchone()

    if not user:
        referred_by = None
        if args and args[0].isdigit() and int(args[0]) != user_id:
            referred_by = int(args[0])
        cursor.execute("INSERT INTO users (user_id, username, referred_by) VALUES (?, ?, ?)", (user_id, username, referred_by))
        conn.commit()
    else:
        cursor.execute("UPDATE users SET username = ? WHERE user_id = ?", (username, user_id))
        conn.commit()

    ref_link = f"https://t.me/{BOT_USERNAME}?start={user_id}"
    total_given = get_total_distributed()
    is_admin = (user_id == ADMIN_ID)

    text = (
        "🚀 **Welcome to Kognitron AI ($KGN) Official Referral Bot**\n\n"
        f"🔗 **Your Unique Referral Link:**\n`{ref_link}`\n\n"
        f"🛒 **Direct Presale Link:**\n[Click here to buy on Smithii]({PRESALE_URL})\n\n"
        "💰 **Reward Rules & Campaign Limits:**\n"
        "• Earn **1 $KGN** for every **$5** purchased by your referrals.\n"
        f"• **Individual Cap:** Max **{int(USER_MAX_CAP)} $KGN** ($200 value at launch) per user.\n"
        f"• **Campaign Pool:** **{total_given:,.0f} / {TOTAL_POOL_CAP:,.0f} $KGN** allocated.\n"
        "• Listing Price: **$1.00** | Presale Price: **$0.80**\n\n"
        "📌 **Quick Commands:**\n"
        "• `/wallet <address>` - Register your Solana wallet for airdrop\n"
        "• `/balance` - View your accumulated $KGN rewards\n"
        "• `/leaderboard` - View top 10 referrers\n"
        "• `/links` - View official verified channels & website\n"
        "• `/about` - Ecosystem overview and rules"
    )
    await update.message.reply_text(text, reply_markup=get_official_links_keyboard(is_admin=is_admin), parse_mode="Markdown", disable_web_page_preview=True)

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "export_data":
        if query.from_user.id != ADMIN_ID:
            await query.message.reply_text("⛔ Unauthorized access.")
            return
        await generate_and_send_export(query.message.chat_id, context)
    
    elif query.data == "show_leaderboard":
        leaderboard_text = get_leaderboard_text()
        await query.message.reply_text(leaderboard_text, parse_mode="Markdown")

async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    await generate_and_send_export(update.effective_chat.id, context)

async def leaderboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    leaderboard_text = get_leaderboard_text()
    await update.message.reply_text(leaderboard_text, parse_mode="Markdown")

async def about(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_given = get_total_distributed()

    text = (
        "**About Kognitron AI Assistant & Rewards Program**\n\n"
        "I am the official utility and rewards bot for the **Kognitron AI ($KGN)** ecosystem.\n\n"
        "🔹 **Campaign Specifications:**\n"
        f"• **Total Reward Pool:** **500,000 $KGN** (0.5% of 100M Total Supply)\n"
        f"• **Max Cap Per Person:** **{int(USER_MAX_CAP)} $KGN**\n"
        "• **Reward Rate:** 1 $KGN per $5 USD raised through your referral\n"
        f"• **Pool Distributed:** `{total_given:,.1f} / {TOTAL_POOL_CAP:,.0f} $KGN`\n\n"
    )

    keyboard = [
        [InlineKeyboardButton("💬 Open Private Chat & Get Link", url="https://t.me/KGN_ref_bot?start=help")],
        [InlineKeyboardButton("🛒 Buy Presale (Smithii)", url=PRESALE_URL)],
        [InlineKeyboardButton("🏆 Leaderboard (Top 10)", callback_data="show_leaderboard")],
        [
            InlineKeyboardButton("🌐 Website", url=WEBSITE_URL),
            InlineKeyboardButton("🐦 Project X", url=TWITTER_PROJECT)
        ],
        [
            InlineKeyboardButton("📢 Announcements", url=ANNOUNCEMENTS_CHANNEL),
            InlineKeyboardButton("💬 Telegram Chat", url=COMMUNITY_CHAT)
        ],
        [
            InlineKeyboardButton("👾 Discord", url=DISCORD_URL),
            InlineKeyboardButton("👤 Founder X", url=TWITTER_FOUNDER)
        ]
    ]

    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown", disable_web_page_preview=True)

async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        await update.message.reply_text("🔒 Please check your balance in private DM.")
        return

    user_id = update.effective_user.id
    cursor.execute("SELECT kgn_balance, wallet_address FROM users WHERE user_id = ?", (user_id,))
    row = cursor.fetchone()

    kgn_bal = row[0] if row else 0.0
    wallet_addr = row[1] if row and row[1] else "Not configured (/wallet)"
    remaining_cap = max(0.0, USER_MAX_CAP - kgn_bal)

    text = (
        "📊 **Your Earnings Dashboard**\n\n"
        f"• **Earned Balance:** `{kgn_bal:.1f} / {int(USER_MAX_CAP)} $KGN`\n"
        f"• **Remaining Cap:** `{remaining_cap:.1f} $KGN`\n"
        f"• **Payout Wallet:** `{wallet_addr}`\n\n"
        "ℹ️ *Rewards will be automatically airdropped to your wallet after presale concludes.*"
    )
    await update.message.reply_text(text, parse_mode="Markdown")

async def links(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "🌐 **Kognitron AI Verified Resources**\n\n"
        f"• **Presale Launch:** [Smithii Launchpad]({PRESALE_URL})\n"
        f"• **Official Website:** {WEBSITE_URL}\n"
        f"• **Contact Email:** `{CONTACT_EMAIL}`\n"
        f"• **Founder:** [Monzer]({TWITTER_FOUNDER})\n\n"
        "Use the buttons below to access our direct channels:"
    )
    await update.message.reply_text(text, reply_markup=get_official_links_keyboard(), parse_mode="Markdown", disable_web_page_preview=True)

async def presale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "🚀 **Kognitron AI ($KGN) Presale is LIVE!**\n\n"
        f"• **Presale Link:** [Buy on Smithii Launchpad]({PRESALE_URL})\n"
        "• **Presale Price:** $0.80\n"
        "• **Listing Price:** $1.00\n"
        "• **Network:** Solana (SOL)"
    )
    await update.message.reply_text(text, reply_markup=get_official_links_keyboard(), parse_mode="Markdown", disable_web_page_preview=True)

async def wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        await update.message.reply_text("🔒 For security, please register your wallet in a private DM with the bot.")
        return

    user_id = update.effective_user.id
    if not context.args:
        await update.message.reply_text(
            "⚠️ Please include your Solana address.\n\n**Format:** `/wallet YourSolanaAddressHere`",
            parse_mode="Markdown"
        )
        return

    sol_address = context.args[0]
    cursor.execute("UPDATE users SET wallet_address = ? WHERE user_id = ?", (sol_address, user_id))
    conn.commit()
    await update.message.reply_text("✅ **Solana address saved successfully.**", parse_mode="Markdown")

async def confirm_buy(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    if len(context.args) < 2:
        await update.message.reply_text("Usage: `/confirm_buy BUYER_USER_ID USD_AMOUNT`", parse_mode="Markdown")
        return

    buyer_id = int(context.args[0])
    usd_amount = float(context.args[1])

    if usd_amount < 5:
        await update.message.reply_text("⚠️ Minimum calculation threshold is $5.")
        return

    total_distributed = get_total_distributed()
    if total_distributed >= TOTAL_POOL_CAP:
        await update.message.reply_text("⚠️ Campaign pool of 500,000 $KGN is completely exhausted.")
        return

    earned_kgn = int(usd_amount // 5)

    cursor.execute("SELECT referred_by FROM users WHERE user_id = ?", (buyer_id,))
    row = cursor.fetchone()

    if row and row[0]:
        referrer_id = row[0]
        cursor.execute("SELECT kgn_balance FROM users WHERE user_id = ?", (referrer_id,))
        ref_row = cursor.fetchone()
        current_bal = ref_row[0] if ref_row else 0.0

        if current_bal >= USER_MAX_CAP:
            await update.message.reply_text(f"ℹ️ Referrer ({referrer_id}) already reached the individual max cap of {USER_MAX_CAP} $KGN.")
            return

        allowed_for_user = USER_MAX_CAP - current_bal
        allowed_from_pool = TOTAL_POOL_CAP - total_distributed
        actual_add = min(earned_kgn, allowed_for_user, allowed_from_pool)

        cursor.execute("UPDATE users SET kgn_balance = kgn_balance + ? WHERE user_id = ?", (actual_add, referrer_id))
        conn.commit()

        new_user_bal = current_bal + actual_add
        new_total_pool = total_distributed + actual_add

        try:
            await context.bot.send_message(
                chat_id=referrer_id,
                text=(
                    f"🎉 **Referral Reward Credited!**\n\n"
                    f"A user you referred completed a **${usd_amount}** purchase.\n"
                    f"**+{actual_add} $KGN** added to your balance.\n"
                    f"Your Balance: **{new_user_bal:.1f} / {int(USER_MAX_CAP)} $KGN**"
                ),
                parse_mode="Markdown"
            )
        except Exception:
            pass

        await update.message.reply_text(
            f"✅ **Success!**\n"
            f"• Credited: `{actual_add} $KGN` to user `{referrer_id}`\n"
            f"• User Total: `{new_user_bal:.1f} / {int(USER_MAX_CAP)} $KGN`\n"
            f"• Global Pool: `{new_total_pool:,.1f} / {TOTAL_POOL_CAP:,.0f} $KGN`",
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text("⚠️ No referrer recorded for this buyer.")

if __name__ == "__main__":
    web_thread = threading.Thread(target=run_web)
    web_thread.daemon = True
    web_thread.start()

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("about", about))
    app.add_handler(CommandHandler("help", about))
    app.add_handler(CommandHandler("links", links))
    app.add_handler(CommandHandler("presale", presale))
    app.add_handler(CommandHandler("wallet", wallet))
    app.add_handler(CommandHandler("balance", balance))
    app.add_handler(CommandHandler("leaderboard", leaderboard_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("confirm_buy", confirm_buy))
    app.add_handler(CallbackQueryHandler(handle_callback))
    print("Bot is live...")
    app.run_polling()
